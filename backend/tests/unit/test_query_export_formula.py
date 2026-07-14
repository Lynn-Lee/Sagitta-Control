"""SAG-011：CSV/XLSX 导出公式注入中和。"""
import csv
from io import BytesIO, StringIO

from openpyxl import load_workbook

from app.routers.query import _build_query_export_file, _neutralize_formula


def test_neutralize_formula_prefixes_dangerous_cells():
    for dangerous in ("=1+1", "+1", "-1", "@cmd", "\t=x", "\rx"):
        out = _neutralize_formula(dangerous)
        assert out.startswith("'"), dangerous
    # 正常值不改动
    assert _neutralize_formula("hello") == "hello"
    assert _neutralize_formula("123") == "123"
    assert _neutralize_formula(42) == 42


def test_csv_export_neutralizes_formula_cells():
    result = {"column_list": ["name"], "rows": [["=HYPERLINK(1)"], ["safe"]]}
    content, _mime, _fname = _build_query_export_file(result, "csv")
    text = content.decode("utf-8-sig")
    reader = list(csv.reader(StringIO(text)))
    data_rows = reader[1:]  # 跳过表头
    assert data_rows[0][1] == "'=HYPERLINK(1)"
    assert data_rows[1][1] == "safe"


def test_xlsx_export_neutralizes_formula_cells():
    result = {"column_list": ["name"], "rows": [["=cmd|' /c calc'!A1"]]}
    content, _mime, _fname = _build_query_export_file(result, "xlsx")
    wb = load_workbook(BytesIO(content))
    ws = wb["QueryResult"]
    # 第 2 行第 2 列（row_num 之后）为数据单元格
    assert ws.cell(row=2, column=2).value == "'=cmd|' /c calc'!A1"
